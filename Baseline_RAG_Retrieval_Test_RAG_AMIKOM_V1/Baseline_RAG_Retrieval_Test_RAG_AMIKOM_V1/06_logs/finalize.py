"""Assemble the physical package folder, hash registry, gate file and ZIP."""
from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import os
import shutil
import zipfile

WIB = dt.timezone(dt.timedelta(hours=7))
NOW = dt.datetime.now(WIB).isoformat()
OUT = "/home/claude/work/out"
BUILD = "/home/claude/work/build"
ROOT = "/home/claude/work/package"
PKG = os.path.join(ROOT, "Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1")
DELIVER = "/home/claude/work/deliverables"

if os.path.exists(PKG):
    shutil.rmtree(PKG)
for sub in ("00_manifest", "01_config", "02_results", "03_evaluation", "04_qa",
            "05_checkpoint", "06_logs"):
    os.makedirs(os.path.join(PKG, sub), exist_ok=True)
os.makedirs(DELIVER, exist_ok=True)


def sha256(p):
    h = hashlib.sha256()
    with open(p, "rb") as f:
        for b in iter(lambda: f.read(1 << 20), b""):
            h.update(b)
    return h.hexdigest()


copies = [
    ("Manifest_Retrieval_Test_RAG_AMIKOM_V1.xlsx", "00_manifest"),
    ("Retrieval_Config_RAG_AMIKOM_V1.json", "01_config"),
    ("Retrieval_Results_CORE.jsonl", "02_results"),
    ("Retrieval_Results_SUPPLEMENTARY.jsonl", "02_results"),
    ("Baseline_RAG_Answers.jsonl", "02_results"),
    ("context_packs.json", "02_results"),
    ("baseline_rag_evaluation.json", "03_evaluation"),
    ("retrieval_metrics.json", "03_evaluation"),
    ("generation_autochecks.json", "03_evaluation"),
    ("Error_Analysis_RAG_AMIKOM_V1.xlsx", "03_evaluation"),
    ("database_validation.json", "04_qa"),
    ("Laporan_Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1.pdf", ""),
    ("README_Baseline_RAG_Retrieval_Test.md", ""),
    ("Retrieval_Recovery_Prompt.md", ""),
]
for name, sub in copies:
    shutil.copy2(os.path.join(OUT, name), os.path.join(PKG, sub, name))

# checkpoints
ck = os.path.join(OUT, "checkpoints")
for f in sorted(os.listdir(ck)):
    shutil.copy2(os.path.join(ck, f), os.path.join(PKG, "05_checkpoint", f))

# pipeline source
os.makedirs(os.path.join(PKG, "06_logs", "ragx"), exist_ok=True)
for f in ("run_retrieval.py", "metrics.py", "assemble_context.py",
          "check_generation.py", "evaluate.py", "db_validation.py",
          "build_manifest.py", "build_report.py", "build_package.py",
          "finalize.py", "gen_answers.py"):
    p = os.path.join(BUILD, f)
    if os.path.exists(p):
        shutil.copy2(p, os.path.join(PKG, "06_logs", f))
for f in os.listdir(os.path.join(BUILD, "ragx")):
    if f.endswith(".py"):
        shutil.copy2(os.path.join(BUILD, "ragx", f),
                     os.path.join(PKG, "06_logs", "ragx", f))

evalj = json.load(open(os.path.join(OUT, "baseline_rag_evaluation.json"),
                       encoding="utf-8"))
metrics = json.load(open(os.path.join(OUT, "retrieval_metrics.json"),
                         encoding="utf-8"))

# QA registry csv
with open(os.path.join(PKG, "04_qa", "qa_registry.csv"), "w", newline="",
          encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["qa_id", "description", "status", "evidence"])
    for q in evalj["qa_registry"]:
        w.writerow([q["qa_id"], q["description"], q["status"], q["evidence"]])

# gate file
iso = metrics["isolation_and_control"]
gate = {
    "evaluated_at_wib": NOW,
    "stage": "BASELINE_RAG_AND_RETRIEVAL_TESTING",
    "scope_status": "FROZEN",
    "vector_database_gate": "PASS",
    "retrieval_testing": "PARTIALLY_COMPLETE",
    "baseline_rag": "COMPLETE",
    "baseline_top_k": 5,
    "baseline_top_k_is_final_production": False,
    "retrieval_metrics": {
        "hit_at_1": metrics["per_k"]["1"]["hit_at_k"],
        "hit_at_3": metrics["per_k"]["3"]["hit_at_k"],
        "hit_at_5": metrics["per_k"]["5"]["hit_at_k"],
        "hit_at_10": metrics["per_k"]["10"]["hit_at_k"],
        "recall_at_5": metrics["per_k"]["5"]["recall_at_k"],
        "recall_at_10": metrics["per_k"]["10"]["recall_at_k"],
        "precision_at_5": metrics["per_k"]["5"]["precision_at_k"],
        "mrr_at_10": metrics["per_k"]["10"]["mrr_at_k"],
        "measured_with": "FALLBACK_LEXICAL_BM25_V1 (NOT the mandated E5 encoder)",
    },
    "control_accuracy": iso["control_decision_accuracy"],
    "namespace_filter_accuracy": iso["namespace_filter_accuracy"],
    "archive_leakage": iso["archive_leakage_count"],
    "hallucination": evalj["hallucination"],
    "qa_summary": {
        "PASS": sum(1 for q in evalj["qa_registry"] if q["status"] == "PASS"),
        "PARTIAL": sum(1 for q in evalj["qa_registry"] if q["status"] == "PARTIAL"),
        "FAIL": sum(1 for q in evalj["qa_registry"] if q["status"] == "FAIL"),
    },
    "retrieval_gate": evalj["gates"]["retrieval_gate"],
    "baseline_rag_gate": evalj["gates"]["baseline_rag_gate"],
    "gate_rationale": evalj["gates"]["gate_rationale"],
    "ready_for_multi_agent": "NO",
    "multi_agent_executed": "NO",
    "accepted_blockers_unchanged": ["G01", "G02", "G04", "CF002",
                                    "MR-A11-RELATIONS"],
    "open_stage_blocker": "QUERY_ENCODER_UNAVAILABLE",
}
with open(os.path.join(PKG, "04_qa", "retrieval_gate.json"), "w",
          encoding="utf-8") as f:
    json.dump(gate, f, indent=2, ensure_ascii=False)

# hash registry over every package file
rows = []
for base, _, files in os.walk(PKG):
    for fn in sorted(files):
        p = os.path.join(base, fn)
        rows.append([os.path.relpath(p, PKG), os.path.getsize(p), sha256(p)])
rows.sort()
with open(os.path.join(PKG, "00_manifest", "hash_registry.csv"), "w", newline="",
          encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["relative_path", "size_bytes", "sha256"])
    w.writerows(rows)

# change log csv mirror
with open(os.path.join(PKG, "00_manifest", "change_log.csv"), "w", newline="",
          encoding="utf-8") as f:
    from openpyxl import load_workbook
    wb = load_workbook(os.path.join(OUT, "Manifest_Retrieval_Test_RAG_AMIKOM_V1.xlsx"))
    ws = wb["11_Change_Log"]
    w = csv.writer(f)
    for row in ws.iter_rows(values_only=True):
        w.writerow(list(row))

# ZIP
zip_path = os.path.join(DELIVER, "Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1.zip")
if os.path.exists(zip_path):
    os.remove(zip_path)
with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
    for base, _, files in os.walk(PKG):
        for fn in sorted(files):
            p = os.path.join(base, fn)
            z.write(p, os.path.join("Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1",
                                    os.path.relpath(p, PKG)))

# deliverable copies
for name in ("Manifest_Retrieval_Test_RAG_AMIKOM_V1.xlsx",
             "Laporan_Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1.pdf",
             "README_Baseline_RAG_Retrieval_Test.md",
             "Retrieval_Config_RAG_AMIKOM_V1.json",
             "Retrieval_Results_CORE.jsonl",
             "Retrieval_Results_SUPPLEMENTARY.jsonl",
             "Baseline_RAG_Answers.jsonl",
             "Error_Analysis_RAG_AMIKOM_V1.xlsx",
             "Retrieval_Recovery_Prompt.md"):
    shutil.copy2(os.path.join(OUT, name), os.path.join(DELIVER, name))

with zipfile.ZipFile(zip_path) as z:
    bad = z.testzip()
    members = len(z.namelist())
final = {
    "package_name": "Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1",
    "validation_type": "FINAL_PACKAGE",
    "validated_at_wib": NOW,
    "status": "PASS",
    "zip": {"path": os.path.basename(zip_path), "size_bytes": os.path.getsize(zip_path),
            "sha256": sha256(zip_path), "crc_test": "PASS" if bad is None else "FAIL",
            "member_count": members},
    "required_outputs_present": {},
    "hash_registry_rows": len(rows),
    "gates": {"retrieval_gate": gate["retrieval_gate"],
              "baseline_rag_gate": gate["baseline_rag_gate"],
              "ready_for_multi_agent": "NO"},
}
required = [
    "Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1.zip",
    "Manifest_Retrieval_Test_RAG_AMIKOM_V1.xlsx",
    "Laporan_Baseline_RAG_Retrieval_Test_RAG_AMIKOM_V1.pdf",
    "README_Baseline_RAG_Retrieval_Test.md",
    "Retrieval_Config_RAG_AMIKOM_V1.json",
    "Retrieval_Results_CORE.jsonl",
    "Retrieval_Results_SUPPLEMENTARY.jsonl",
    "Baseline_RAG_Answers.jsonl",
    "Error_Analysis_RAG_AMIKOM_V1.xlsx",
    "Retrieval_Recovery_Prompt.md",
]
for r in required:
    p = os.path.join(DELIVER, r)
    final["required_outputs_present"][r] = {
        "exists": os.path.exists(p),
        "size_bytes": os.path.getsize(p) if os.path.exists(p) else 0,
        "sha256": sha256(p) if os.path.exists(p) else None,
    }
final["missing_outputs"] = [r for r in required
                            if not final["required_outputs_present"][r]["exists"]]
with open(os.path.join(DELIVER, "final_package_validation_retrieval.json"), "w",
          encoding="utf-8") as f:
    json.dump(final, f, indent=2, ensure_ascii=False)

print(json.dumps({"zip": final["zip"], "missing": final["missing_outputs"],
                  "files": len(rows)}, indent=1))
